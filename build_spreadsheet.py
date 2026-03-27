"""
Build the output spreadsheet with all holiday options, rankings, and flight grids.
"""

from __future__ import annotations

import os
from datetime import date
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

import config


# ── Colours ───────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TOP10_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ESTIMATE_FONT = Font(italic=True, color="808080")
LINK_FONT = Font(color="0563C1", underline="single")
GBP_FORMAT = '£#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def build_spreadsheet(enriched_deals: list[dict], api_log: list[dict]) -> str:
    """Create the output XLSX with 4 sheets. Returns the file path."""
    wb = Workbook()

    # Sheet 1: All Options
    ws1 = wb.active
    ws1.title = "All Options"
    _build_all_options(ws1, enriched_deals)

    # Sheet 2: Top 10
    ws2 = wb.create_sheet("Top 10 Cheapest")
    _build_all_options(ws2, enriched_deals[:10], is_top10=True)

    # Sheet 3: Flight Price Grid
    ws3 = wb.create_sheet("Flight Price Grid")
    _build_flight_grid(ws3, enriched_deals)

    # Sheet 4: Raw Data Log
    ws4 = wb.create_sheet("Raw Data Log")
    _build_raw_log(ws4, api_log)

    # Save
    os.makedirs(os.path.dirname(config.OUTPUT_PATH), exist_ok=True)
    wb.save(config.OUTPUT_PATH)
    return config.OUTPUT_PATH


# ── Sheet 1 & 2: All Options / Top 10 ────────────────────────

COLUMNS = [
    "Rank",
    "TOTAL £",
    "Pickup City",
    "Dropoff City",
    "Drive Hours",
    "Pickup Date",
    "Drive Days",
    "Vehicle",
    "Imoova £",
    "Outbound Date",
    "Outbound Airline",
    "Outbound Route",
    "Outbound £",
    "Return Date",
    "Return Airline",
    "Return Route",
    "Return £",
    "UK Transport",
    "UK Transport £",
    "Complete?",
    "Warnings",
    "Imoova Link",
    "Outbound Flight Link",
    "Return Flight Link",
]


def _build_all_options(ws, deals: list[dict], is_top10: bool = False):
    """Build the main options table."""
    # Header row
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, item in enumerate(deals, 2):
        deal = item["deal"]
        ob = item.get("cheapest_outbound")
        ret = item.get("cheapest_return")
        ob_uk = item.get("outbound_uk_transport")
        ret_uk = item.get("return_uk_transport")

        # UK transport details
        uk_detail = ""
        uk_cost = 0.0
        if ob_uk:
            uk_detail = ob_uk.get("details", "")
            uk_cost += ob_uk.get("price_gbp", 0)
        if ret_uk:
            if uk_detail:
                uk_detail += " + "
            uk_detail += ret_uk.get("details", "")
            uk_cost += ret_uk.get("price_gbp", 0)

        # Google Flights deeplinks — one-way outbound TO pickup, one-way return FROM dropoff
        gf_outbound_link = ""
        if ob:
            gf_outbound_link = _google_flights_link(
                ob.get("departure_airport", ""),
                ob.get("arrival_airport", ""),
                ob.get("search_date", deal["depart_date"]),
            )
        elif not config.is_london(deal["pickup_city"]):
            pickup_airports = config.get_airports_for_city(deal["pickup_city"])
            if pickup_airports:
                gf_outbound_link = _google_flights_link("LON", pickup_airports[0], deal["depart_date"])

        gf_return_link = ""
        if ret:
            gf_return_link = _google_flights_link(
                ret.get("departure_airport", ""),
                ret.get("arrival_airport", ""),
                ret.get("search_date", deal["deliver_date"]),
            )
        elif not config.is_london(deal["dropoff_city"]):
            dropoff_airports = config.get_airports_for_city(deal["dropoff_city"])
            if dropoff_airports:
                gf_return_link = _google_flights_link(dropoff_airports[0], "LON", deal["deliver_date"])

        drive_hours = config.estimate_driving_hours(deal["pickup_city"], deal["dropoff_city"])

        values = [
            f"=ROW()-1",
            item["total_cost"],
            deal["pickup_city"],
            deal["dropoff_city"],
            f"~{drive_hours}h" if drive_hours else "?",
            deal["depart_date"],
            deal["drive_days"],
            deal["vehicle"],
            deal["rate_gbp"] / 2,  # per-person: split Imoova fee with partner
            ob.get("search_date", "") if ob else "",
            ob.get("airline", "") if ob else "",
            f"{ob['departure_airport']}->{ob['arrival_airport']}" if ob else "",
            ob.get("price_gbp", "") if ob else "",
            ret.get("search_date", "") if ret else "",
            ret.get("airline", "") if ret else "",
            f"{ret['departure_airport']}->{ret['arrival_airport']}" if ret else "",
            ret.get("price_gbp", "") if ret else "",
            uk_detail,
            uk_cost if uk_cost > 0 else "",
            "Yes" if item.get("is_complete") else "No",
            "; ".join(item.get("warnings", [])),
            deal.get("deal_url", ""),
            gf_outbound_link,
            gf_return_link,
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # GBP formatting for price columns
            if COLUMNS[col_idx - 1].endswith("£"):
                if isinstance(value, (int, float)) and value != "":
                    cell.number_format = GBP_FORMAT

        # ── Conditional formatting ────────────────────────────
        total_cell = ws.cell(row=row_idx, column=COLUMNS.index("TOTAL £") + 1)
        total_val = item["total_cost"]

        if total_val < 80:
            total_cell.fill = GREEN_FILL
        elif total_val < 150:
            total_cell.fill = YELLOW_FILL
        elif total_val < 9999:
            total_cell.fill = RED_FILL

        # Hyperlinks
        deal_url = deal.get("deal_url", "")
        if deal_url:
            link_cell = ws.cell(row=row_idx, column=COLUMNS.index("Imoova Link") + 1)
            link_cell.hyperlink = deal_url
            link_cell.font = LINK_FONT

        if gf_outbound_link:
            gf_out_cell = ws.cell(row=row_idx, column=COLUMNS.index("Outbound Flight Link") + 1)
            gf_out_cell.hyperlink = gf_outbound_link
            gf_out_cell.font = LINK_FONT

        if gf_return_link:
            gf_ret_cell = ws.cell(row=row_idx, column=COLUMNS.index("Return Flight Link") + 1)
            gf_ret_cell.hyperlink = gf_return_link
            gf_ret_cell.font = LINK_FONT

        # Estimate styling
        if ob_uk and ob_uk.get("is_estimate"):
            uk_cell = ws.cell(row=row_idx, column=COLUMNS.index("UK Transport") + 1)
            uk_cell.font = ESTIMATE_FONT
        if ret_uk and ret_uk.get("is_estimate"):
            uk_cell = ws.cell(row=row_idx, column=COLUMNS.index("UK Transport") + 1)
            uk_cell.font = ESTIMATE_FONT

    # Top 10 highlight via Excel conditional formatting (dynamic)
    if not is_top10:
        last_col = get_column_letter(len(COLUMNS))
        last_row = len(deals) + 1
        data_range = f"A2:{last_col}{last_row}"
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=["$A2<=10"],
                fill=TOP10_FILL,
            ),
        )

    # Auto-fit column widths (approximate)
    _autofit_columns(ws)


# ── Sheet 3: Flight Price Grid ────────────────────────────────

def _build_flight_grid(ws, enriched_deals: list[dict]):
    """Build two grids: London->City and City->London cheapest prices by date."""

    # Collect all unique cities and dates from flight results
    outbound_prices = defaultdict(dict)  # city -> date -> price
    return_prices = defaultdict(dict)

    for item in enriched_deals:
        deal = item["deal"]
        pickup = deal["pickup_city"]
        dropoff = deal["dropoff_city"]

        for flight in item.get("outbound_flights", []):
            d = flight.get("search_date", "")
            p = flight.get("price_gbp", 0)
            city = pickup
            if d and p:
                existing = outbound_prices[city].get(d, 9999)
                outbound_prices[city][d] = min(existing, p)

        for flight in item.get("return_flights", []):
            d = flight.get("search_date", "")
            p = flight.get("price_gbp", 0)
            city = dropoff
            if d and p:
                existing = return_prices[city].get(d, 9999)
                return_prices[city][d] = min(existing, p)

    # Get sorted unique dates
    all_out_dates = sorted(set(d for city_dates in outbound_prices.values() for d in city_dates))
    all_ret_dates = sorted(set(d for city_dates in return_prices.values() for d in city_dates))

    # ── Grid 1: London -> City ────────────────────────────────
    row = 1
    ws.cell(row=row, column=1, value="LONDON -> CITY (Outbound)").font = Font(bold=True, size=13)
    row += 1

    # Headers
    ws.cell(row=row, column=1, value="City").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for j, d in enumerate(all_out_dates, 2):
        cell = ws.cell(row=row, column=j, value=d)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for city in sorted(outbound_prices.keys()):
        ws.cell(row=row, column=1, value=city).font = Font(bold=True)
        for j, d in enumerate(all_out_dates, 2):
            price = outbound_prices[city].get(d)
            if price and price < 9999:
                cell = ws.cell(row=row, column=j, value=price)
                cell.number_format = GBP_FORMAT
                if price < 30:
                    cell.fill = GREEN_FILL
                elif price < 60:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = RED_FILL
        row += 1

    # ── Grid 2: City -> London ────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="CITY -> LONDON (Return)").font = Font(bold=True, size=13)
    row += 1

    ws.cell(row=row, column=1, value="City").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for j, d in enumerate(all_ret_dates, 2):
        cell = ws.cell(row=row, column=j, value=d)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for city in sorted(return_prices.keys()):
        ws.cell(row=row, column=1, value=city).font = Font(bold=True)
        for j, d in enumerate(all_ret_dates, 2):
            price = return_prices[city].get(d)
            if price and price < 9999:
                cell = ws.cell(row=row, column=j, value=price)
                cell.number_format = GBP_FORMAT
                if price < 30:
                    cell.fill = GREEN_FILL
                elif price < 60:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = RED_FILL
        row += 1

    _autofit_columns(ws)


# ── Sheet 4: Raw Data Log ────────────────────────────────────

def _build_raw_log(ws, api_log: list[dict]):
    """Write every API call as a row."""
    headers = ["Timestamp", "Provider", "From", "To", "Date", "Status", "Results"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    ws.freeze_panes = "A2"

    for row_idx, entry in enumerate(api_log, 2):
        params = entry.get("params", {})
        ws.cell(row=row_idx, column=1, value=entry.get("timestamp", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("provider", ""))
        ws.cell(row=row_idx, column=3, value=params.get("from", ""))
        ws.cell(row=row_idx, column=4, value=params.get("to", ""))
        ws.cell(row=row_idx, column=5, value=params.get("date", ""))
        ws.cell(row=row_idx, column=6, value=entry.get("status", ""))
        ws.cell(row=row_idx, column=7, value=entry.get("result_count", 0))

    _autofit_columns(ws)


# ── Helpers ───────────────────────────────────────────────────

def _google_flights_link(from_code: str, to_code: str, travel_date: str) -> str:
    """Build a Google Flights deeplink for manual verification."""
    return (
        f"https://www.google.com/travel/flights?q="
        f"Flights+from+{from_code}+to+{to_code}+on+{travel_date}+one+way"
    )


def _autofit_columns(ws, min_width: int = 8, max_width: int = 40):
    """Approximate auto-fit by scanning cell contents."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, min(cell_len + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


if __name__ == "__main__":
    # Test with dummy data
    print("=== Spreadsheet Builder Test ===")
    test_deals = [{
        "deal": {
            "ref": "TEST001",
            "pickup_city": "Barcelona",
            "dropoff_city": "Lyon",
            "depart_date": "2026-03-29",
            "deliver_date": "2026-04-05",
            "drive_days": 7,
            "vehicle": "Fiat Ducato",
            "rate_raw": "€1/day",
            "rate_gbp": 6.02,
            "seats": 2,
            "fuel": "Included",
            "ferry": "",
            "deal_url": "https://www.imoova.com/en/relocations/deal/test",
        },
        "outbound_flights": [],
        "return_flights": [],
        "outbound_uk_transport": None,
        "return_uk_transport": None,
        "cheapest_outbound": {
            "provider": "fast-flights",
            "airline": "Ryanair",
            "departure_airport": "STN",
            "arrival_airport": "BCN",
            "departure_time": "06:15",
            "arrival_time": "09:30",
            "duration": "2h 15m",
            "stops": 0,
            "price_gbp": 19.0,
            "is_best": True,
            "search_date": "2026-03-28",
            "raw_price_str": "£19",
        },
        "cheapest_return": {
            "provider": "fast-flights",
            "airline": "easyJet",
            "departure_airport": "LYS",
            "arrival_airport": "LGW",
            "departure_time": "18:40",
            "arrival_time": "19:50",
            "duration": "1h 50m",
            "stops": 0,
            "price_gbp": 22.0,
            "is_best": False,
            "search_date": "2026-04-05",
            "raw_price_str": "£22",
        },
        "total_cost": 47.02,
        "is_complete": True,
        "warnings": [],
    }]

    path = build_spreadsheet(test_deals, [])
    print(f"Test spreadsheet saved to: {path}")
